import openpyxl
import os


def makeexcelfiles():
    inputworkbook = openpyxl.load_workbook("HomeworkGrade.xlsx")
    sheetnames = inputworkbook.sheetnames
    print(sheetnames)
    worksheet = inputworkbook[sheetnames[1]]

    rows = worksheet.max_row
    columns = worksheet.max_column

    # The name of output directory
    OutputDirName = 'IndividualFiles'

    if not os.path.exists(OutputDirName):
        os.makedirs(OutputDirName)

    # Note that the filename is automatically taken as the content of first column
    # Which I set to student ID

    for i in range(2, rows + 1):
        outputworkbook = openpyxl.Workbook()
        outputworksheet = outputworkbook.active
        outputworksheet.title = "Sheet1"
        for j in range(1, columns + 1):
            outputworksheet.cell(1, j, worksheet.cell(row=1, column=j).value)
            outputworksheet.cell(2, j, worksheet.cell(row=i, column=j).value)
        outputfilename = OutputDirName + "/" + worksheet.cell(
            row=i, column=1).value + ".xlsx"
        outputworkbook.save(filename=outputfilename)
    return


if __name__ == "__main__":
    makeexcelfiles()
