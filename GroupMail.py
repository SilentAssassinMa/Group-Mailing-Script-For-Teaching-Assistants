#########################
##  Group Mail Script  ##
##  Auther : Wenhao Ma ##
#########################

import openpyxl
from IndividualFileGenerate import makeexcelfiles
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

# If you want to generate individual files (e.g. Homework Grade),
# set the value to true

doGenerateIndividualFiles = True

if (doGenerateIndividualFiles):
    makeexcelfiles()

mail_host = 'mail.ustc.edu.cn'  # SMTP server
mail_user = 'silentassassin@mail.ustc.edu.cn'  # Your Email Address
mail_pass = '123456'  # Your Password
sender = 'silentassassin@mail.ustc.edu.cn'  # The Sender Your Want to Show

inputworkbook = openpyxl.load_workbook("MailList.xlsx", data_only=True)
sheetnames = inputworkbook.sheetnames
worksheet = inputworkbook[sheetnames[0]]
rows = worksheet.max_row
columns = worksheet.max_column

mainbody = '您好, 您本次期末考试的成绩为 :'
count = 0
for i in range(1, rows + 1):
    content = 'Dear %s %s : \n\n' % (worksheet.cell(
        row=i, column=1).value, worksheet.cell(
            row=i, column=2).value) + mainbody + '\n'

    # You may change the following line to suit your needs
    content += '  填空题: %s \n' % (worksheet.cell(row=i, column=4).value)
    content += '  判断题: %s \n' % (worksheet.cell(row=i, column=5).value)
    content += '  解答题1: %s \n' % (worksheet.cell(row=i, column=6).value)
    content += '  解答题2: %s \n' % (worksheet.cell(row=i, column=7).value)
    content += '  解答题3: %s \n' % (worksheet.cell(row=i, column=8).value)
    content += '  解答题4: %s \n' % (worksheet.cell(row=i, column=9).value)
    content += '  解答题5: %s \n' % (worksheet.cell(row=i, column=10).value)
    content += '  解答题6: %s \n' % (worksheet.cell(row=i, column=11).value)
    content += '  解答题7: %s \n' % (worksheet.cell(row=i, column=12).value)
    content += '  总分: %s \n' % (worksheet.cell(row=i, column=13).value)

    content += "祝好 \n"
    content += "王五"

    message = MIMEMultipart()
    message['Subject'] = '期末考试成绩'
    message['From'] = sender
    message['To'] = worksheet.cell(row=i, column=3).value

    message.attach(MIMEText(content, 'plain', 'utf-8'))

    if (doGenerateIndividualFiles):
        excelFile = 'IndividualFiles/%s.xlsx' % worksheet.cell(row=i,
                                                               column=1).value
        excelPart = MIMEApplication(open(excelFile, 'rb').read())
        excelPart.add_header('Content-Disposition',
                             'attachment',
                             filename='作业成绩.xlsx')
        message.attach(excelPart)

    receiver = message['To']

    print("send to : ", message['To'])
    smtpObj = smtplib.SMTP()
    smtpObj.connect(mail_host, 25)
    smtpObj.login(mail_user, mail_pass)
    smtpObj.sendmail(sender, receiver, message.as_string())
    smtpObj.quit()
    count += 1
    print(count, ' success')